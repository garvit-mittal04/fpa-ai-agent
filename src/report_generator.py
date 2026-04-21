"""
report_generator.py
-------------------
Generates a polished 3-sheet Excel report:
  Sheet 1 — Executive Summary
  Sheet 2 — Variance Detail
  Sheet 3 — AI Commentary

Upgrades:
- KPI block at top
- Better formatting for currency / percent
- Conditional status colors
- Highlight anomalies and large variances
- Freeze panes and filters
- Cleaner commentary sheet layout
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Theme styles
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=11, color="44546A")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

ALT_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
RED_FILL = PatternFill(start_color="FDE9E7", end_color="FDE9E7", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")

CURRENCY_FORMAT = '$#,##0.00'
PERCENT_FORMAT = '0.00%'


def _style_header_row(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _set_currency(cell):
    cell.number_format = CURRENCY_FORMAT


def _set_percent(cell):
    cell.number_format = PERCENT_FORMAT


def _autosize_columns(ws, min_width=12, max_width=40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _safe_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _status_from_variance_pct(var_pct: float) -> str:
    if abs(var_pct) <= 5:
        return "On Track"
    elif abs(var_pct) <= 10:
        return "Watch"
    return "At Risk"


def _status_fill(status: str):
    if status == "On Track":
        return GREEN_FILL
    if status == "Watch":
        return YELLOW_FILL
    return RED_FILL


def generate_excel_report(
    variance_df: pd.DataFrame,
    dept_df: pd.DataFrame,
    commentary: str,
    output_path: str
):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()

    # =========================
    # Sheet 1: Executive Summary
    # =========================
    ws1 = wb.active
    ws1.title = "Executive Summary"

    total_actual = _safe_float(variance_df.get("actual_amount", pd.Series(dtype=float)).sum())
    total_budget = _safe_float(variance_df.get("budget_amount", pd.Series(dtype=float)).sum())
    total_variance = total_actual - total_budget
    total_variance_pct = (total_variance / total_budget) if total_budget != 0 else 0.0
    anomaly_count = int(variance_df["is_anomaly"].sum()) if "is_anomaly" in variance_df.columns else 0
    report_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws1["A1"] = "FP&A Variance Analysis — Executive Summary"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:F1")

    ws1["A2"] = f"Generated: {report_time}"
    ws1["A2"].font = SUBTITLE_FONT
    ws1.merge_cells("A2:F2")

    # KPI block
    kpi_labels = ["Total Actual", "Total Budget", "Net Variance", "Variance %", "Anomalies"]
    kpi_values = [total_actual, total_budget, total_variance, total_variance_pct, anomaly_count]

    for i, (label, value) in enumerate(zip(kpi_labels, kpi_values), start=1):
        col = i
        label_cell = ws1.cell(row=4, column=col, value=label)
        label_cell.fill = HEADER_FILL
        label_cell.font = HEADER_FONT
        label_cell.alignment = CENTER
        label_cell.border = THIN_BORDER

        value_cell = ws1.cell(row=5, column=col, value=value)
        value_cell.border = THIN_BORDER
        value_cell.alignment = CENTER
        value_cell.fill = GRAY_FILL

        if label in {"Total Actual", "Total Budget", "Net Variance"}:
            _set_currency(value_cell)
        elif label == "Variance %":
            _set_percent(value_cell)

    headers = ["Department", "Total Actual", "Total Budget", "Variance $", "Variance %", "Status"]
    start_row = 8
    _style_header_row(ws1, headers, row=start_row)

    for i, row in dept_df.iterrows():
        r = start_row + 1 + i
        dept = row.iloc[0]
        total_actual_dept = _safe_float(row.iloc[1])
        total_budget_dept = _safe_float(row.iloc[2])
        variance_dollar = _safe_float(row.iloc[3])
        variance_pct_raw = _safe_float(row.iloc[4]) / 100.0  # convert from % value to Excel percent
        status = _status_from_variance_pct(variance_pct_raw * 100)

        values = [
            dept,
            total_actual_dept,
            total_budget_dept,
            variance_dollar,
            variance_pct_raw,
            status,
        ]

        row_fill = ALT_FILL if i % 2 == 0 else PatternFill()

        for col, val in enumerate(values, 1):
            cell = ws1.cell(row=r, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col != 1 else LEFT

            if col in {1, 2, 3, 4, 5}:
                cell.fill = row_fill
            if col == 6:
                cell.fill = _status_fill(status)

            if col in {2, 3, 4}:
                _set_currency(cell)
            elif col == 5:
                _set_percent(cell)

    ws1.freeze_panes = "A9"
    ws1.auto_filter.ref = f"A{start_row}:F{ws1.max_row}"
    _autosize_columns(ws1)

    # =========================
    # Sheet 2: Variance Detail
    # =========================
    ws2 = wb.create_sheet("Variance Detail")

    detail_headers = [
        "Department",
        "Line Item",
        "Period",
        "Actual",
        "Budget",
        "Variance $",
        "Variance %",
        "Anomaly",
    ]
    _style_header_row(ws2, detail_headers, row=1)

    for i, row in variance_df.iterrows():
        r = i + 2
        is_anomaly = bool(row.get("is_anomaly", False))
        variance_pct_raw = _safe_float(row.get("variance_pct", 0)) / 100.0

        values = [
            row.get("department", ""),
            row.get("line_item", ""),
            row.get("period", ""),
            _safe_float(row.get("actual_amount", 0)),
            _safe_float(row.get("budget_amount", 0)),
            _safe_float(row.get("variance_dollar", 0)),
            variance_pct_raw,
            "Yes" if is_anomaly else "No",
        ]

        if is_anomaly:
            row_fill = RED_FILL
        else:
            row_fill = ALT_FILL if i % 2 == 0 else PatternFill()

        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=r, column=col, value=val)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col not in {1, 2} else LEFT

            if col in {4, 5, 6}:
                _set_currency(cell)
            elif col == 7:
                _set_percent(cell)

        # Extra highlight for large variances even if not anomaly
        abs_var_pct = abs(_safe_float(row.get("variance_pct", 0)))
        if abs_var_pct > 10 and not is_anomaly:
            for col in range(1, 9):
                ws2.cell(row=r, column=col).fill = YELLOW_FILL

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:H{ws2.max_row}"
    _autosize_columns(ws2)

    # =========================
    # Sheet 3: AI Commentary
    # =========================
    ws3 = wb.create_sheet("AI Commentary")

    ws3["A1"] = "AI-Generated Management Commentary"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells("A1:G1")

    ws3["A2"] = f"Generated: {report_time}"
    ws3["A2"].font = SUBTITLE_FONT
    ws3.merge_cells("A2:G2")

    ws3["A4"] = "Executive Narrative"
    ws3["A4"].fill = HEADER_FILL
    ws3["A4"].font = HEADER_FONT
    ws3["A4"].alignment = CENTER
    ws3["A4"].border = THIN_BORDER
    ws3.merge_cells("A4:G4")

    ws3["A5"] = commentary
    ws3["A5"].alignment = WRAP
    ws3["A5"].border = THIN_BORDER
    ws3["A5"].fill = GRAY_FILL
    ws3.merge_cells("A5:G28")

    ws3["A30"] = "Summary Metrics"
    ws3["A30"].fill = HEADER_FILL
    ws3["A30"].font = HEADER_FONT
    ws3["A30"].alignment = CENTER
    ws3["A30"].border = THIN_BORDER
    ws3.merge_cells("A30:C30")

    summary_rows = [
        ("Total Actual", total_actual),
        ("Total Budget", total_budget),
        ("Net Variance", total_variance),
        ("Variance %", total_variance_pct),
        ("Anomalies", anomaly_count),
    ]

    row_idx = 31
    for label, value in summary_rows:
        ws3.cell(row=row_idx, column=1, value=label).border = THIN_BORDER
        ws3.cell(row=row_idx, column=1).fill = ALT_FILL
        ws3.cell(row=row_idx, column=1).font = Font(bold=True)

        val_cell = ws3.cell(row=row_idx, column=2, value=value)
        val_cell.border = THIN_BORDER

        if label in {"Total Actual", "Total Budget", "Net Variance"}:
            _set_currency(val_cell)
        elif label == "Variance %":
            _set_percent(val_cell)

        row_idx += 1

    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 18
    ws3.column_dimensions["F"].width = 18
    ws3.column_dimensions["G"].width = 18
    ws3.row_dimensions[5].height = 220

    wb.save(output_path)
